import threading
import os
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from Apps.Administracion.models import *
from Apps.Notificaciones.models import Notificaciones
from Apps.Solicitudes.models import Solicitudes


# Importa tus modelos aquí
# from mis_apps.models import Persona, Area, AreaPersona, Grupo, GrupoPersona, Solicitudes, Notificaciones, CoreChoices

class DjangoReportThreadPool:
    """
    Clase encargada de la generación de reportes en Excel utilizando hilos en Django.
    Al finalizar cada reporte, se genera una notificación en el sistema para el usuario solicitante
    con la URL de descarga del archivo generado.
    """

    def __init__(self, usuario_solicitante, *args, **kwargs):
        self.usuario = usuario_solicitante
        # Buscamos la Persona asociada al usuario para las notificaciones
        try:
            self.persona_solicitante = usuario_solicitante.persona_set.first()  # Ajustar según tu relación inversa real
        except Exception:
            self.persona_solicitante = None

        # Carpeta donde se guardarán los reportes temporal/permanentemente dentro de MEDIA
        self.media_report_path = os.path.join(settings.MEDIA_ROOT, 'reportes_excel')
        if not os.path.exists(self.media_report_path):
            os.makedirs(self.media_report_path, exist_ok=True)

    def _crear_notificacion_exito(self, titulo, descripcion, nombre_archivo):
        """Método auxiliar interno para registrar la notificación final con la URL del Excel"""
        if self.persona_solicitante:
            url_descarga = f"{settings.MEDIA_URL}reportes_excel/{nombre_archivo}"
            # Reemplaza 'Notificaciones' por el nombre exacto de tu modelo si varía
            Notificaciones.objects.create(
                destinatario=self.persona_solicitante,
                titulo=titulo,
                descripcion=descripcion,
                estado_notificacion=False,
                tipo_notificacion='B',
                # Asumiendo BAJA o la prioridad que consideres por defecto (ej: CoreChoices.TipoNotificacion.BAJA)
                url=url_descarga
            )

    def _crear_notificacion_error(self, titulo, error_msg):
        """Método auxiliar interno por si el hilo falla en tiempo de ejecución"""
        if self.persona_solicitante:
            Notificaciones.objects.create(
                destinatario=self.persona_solicitante,
                titulo=f"Error: {titulo}",
                descripcion=f"Ocurrió un inconveniente al procesar tu reporte: {str(error_msg)}",
                estado_notificacion=False,
                tipo_notificacion='A',  # Prioridad ALTA por ser un fallo
                url=""
            )

    # =========================================================================
    # 1. REPORTE DE TODAS LAS PERSONAS CON TODOS SUS DATOS
    # =========================================================================
    def reporte_personas_thread(self, **filtros):
        """
        Ejecuta la generación del reporte de personas en un hilo secundario.
        Acepta filtros en formato de diccionario (**kwargs) para futuras expansiones.
        """

        def _proceso():
            try:
                # 1. Inicializar openpyxl Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Personal"

                # Estilos básicos de presentación
                font_titulo = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                fill_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                alignment_center = Alignment(horizontal="center", vertical="center")

                # Cabeceras del reporte basados en los campos de tu modelo Persona
                headers = [
                    "Identificación", "Nombres Completos", "1er Apellido", "2do Apellido", "Nombres",
                    "Fecha Nacimiento", "Sexo", "País", "Provincia", "Cantón", "Teléfono Móvil", "Email"
                ]
                ws.append(headers)

                # Aplicar estilos a las cabeceras
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = font_titulo
                    cell.fill = fill_titulo
                    cell.alignment = alignment_center

                # 2. Consultar datos aplicando filtros dinámicos (si vienen vacíos trae todos)
                # Ejemplo futuros filtros: filtros = {'pais_id': 1, 'sexo': 'M'}
                queryset = Persona.objects.filter(**filtros)

                # 3. Llenar el Excel
                for persona in queryset:
                    ws.append([
                        persona.identificacion,
                        persona.nombre_completo_minus(),
                        persona.apellido1,
                        persona.apellido2,
                        persona.nombres,
                        persona.nacimiento.strftime('%Y-%m-%d') if persona.nacimiento else "",
                        persona.get_sexo_display() if hasattr(persona, 'get_sexo_display') else persona.sexo,
                        persona.pais.nombre if persona.pais else "",
                        persona.provincia.nombre if persona.provincia else "",
                        persona.canton.nombre if persona.canton else "",
                        persona.telefono,
                        persona.email
                    ])

                # Ajustar el ancho automático de las columnas
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                # 4. Guardar archivo físico en MEDIA
                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                nombre_archivo = f"reporte_personas_{self.usuario.id}_{timestamp}.xlsx"
                filepath = os.path.join(self.media_report_path, nombre_archivo)
                wb.save(filepath)

                # 5. Notificar éxito al usuario
                self._crear_notificacion_exito(
                    titulo="Reporte de Personas Listo",
                    descripcion="Tu reporte de personal solicitado ha sido generado con éxito. Haz clic aquí para descargarlo.",
                    nombre_archivo=nombre_archivo
                )

            except Exception as e:
                self._crear_notificacion_error(titulo="Reporte de Personas", error_msg=e)

        # Lanzar el proceso en segundo plano (hilo)
        threading.Thread(target=_proceso, daemon=True).start()

    # =========================================================================
    # 2. REPORTE DE TODAS LAS ÁREAS REGISTRADAS
    # =========================================================================
    def reporte_areas_thread(self, **filtros):
        """
        Genera un reporte con todas las áreas del sistema y su director respectivo.
        """

        def _proceso():
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Áreas"

                font_titulo = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                fill_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                alignment_center = Alignment(horizontal="center", vertical="center")

                headers = ["Nombre del Área", "Descripción", "Identificación Director", "Nombre del Director",
                           "Email del Director"]
                ws.append(headers)

                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = font_titulo
                    cell.fill = fill_titulo
                    cell.alignment = alignment_center

                queryset = Area.objects.filter(**filtros)

                for area in queryset:
                    director = area.director
                    ws.append([
                        area.nombre,
                        area.descripcion,
                        director.identificacion if director else "N/A",
                        director.nombre_completo_minus() if director else "Sin Director",
                        director.email if director else "N/A"
                    ])

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                nombre_archivo = f"reporte_areas_{self.usuario.id}_{timestamp}.xlsx"
                filepath = os.path.join(self.media_report_path, nombre_archivo)
                wb.save(filepath)

                self._crear_notificacion_exito(
                    titulo="Reporte de Áreas Listo",
                    descripcion="El reporte de Áreas registradas se ha generado. Clic para descargar.",
                    nombre_archivo=nombre_archivo
                )

            except Exception as e:
                self._crear_notificacion_error(titulo="Reporte de Áreas", error_msg=e)

        threading.Thread(target=_proceso, daemon=True).start()

    # =========================================================================
    # 3. REPORTE DE PERSONAS POR ÁREA (PLANTILLAS)
    # =========================================================================
    def reporte_plantilla_area_thread(self, area_id=None, **filtros):
        """
        Genera el reporte de los trabajadores asignados a un área.
        Recibe area_id explícitamente o filtra todos si no se envía.
        """

        def _proceso():
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Plantilla por Área"

                font_titulo = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                fill_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

                headers = ["Área", "Director del Área", "Identificación", "Nombres Completos Trabajador",
                           "Email Trabajador", "Teléfono"]
                ws.append(headers)

                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = font_titulo
                    cell.fill = fill_titulo

                if area_id:
                    filtros['area_id'] = area_id

                queryset = AreaPersona.objects.filter(**filtros).select_related('area', 'persona', 'area__director')

                for ap in queryset:
                    area = ap.area
                    persona = ap.persona
                    if area and persona:
                        director = area.director
                        ws.append([
                            area.nombre,
                            director.nombre_completo_minus() if director else "Sin Director",
                            persona.identificacion,
                            persona.nombre_completo_minus(),
                            persona.email,
                            persona.telefono
                        ])

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                nombre_archivo = f"reporte_plantilla_area_{self.usuario.id}_{timestamp}.xlsx"
                filepath = os.path.join(self.media_report_path, nombre_archivo)
                wb.save(filepath)

                self._crear_notificacion_exito(
                    titulo="Reporte Plantilla de Área Listo",
                    descripcion="El reporte de trabajadores por área se ha generado con éxito.",
                    nombre_archivo=nombre_archivo
                )

            except Exception as e:
                self._crear_notificacion_error(titulo="Reporte Plantilla de Área", error_msg=e)

        threading.Thread(target=_proceso, daemon=True).start()

    # =========================================================================
    # 4A. REPORTE GENERAL DE GRUPOS
    # =========================================================================
    def reporte_grupos_thread(self, **filtros):
        """
        Genera el listado de todos los grupos.
        """

        def _proceso():
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Grupos"

                font_titulo = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                fill_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

                headers = ["ID Grupo", "Nombre del Grupo", "Descripción"]
                ws.append(headers)

                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = font_titulo
                    cell.fill = fill_titulo

                queryset = Grupo.objects.filter(**filtros)
                for grupo in queryset:
                    ws.append([grupo.id, grupo.nombre, grupo.descripcion])

                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 30

                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                nombre_archivo = f"reporte_grupos_{self.usuario.id}_{timestamp}.xlsx"
                filepath = os.path.join(self.media_report_path, nombre_archivo)
                wb.save(filepath)

                self._crear_notificacion_exito(
                    titulo="Reporte de Grupos Listo",
                    descripcion="El listado general de grupos ya está disponible.",
                    nombre_archivo=nombre_archivo
                )

            except Exception as e:
                self._crear_notificacion_error(titulo="Reporte de Grupos", error_msg=e)

        threading.Thread(target=_proceso, daemon=True).start()

    # =========================================================================
    # 4B. REPORTE DE PERSONAS EN UN GRUPO (RECIBE PARAMETRO GRUPO_ID)
    # =========================================================================
    def reporte_integrantes_grupo_thread(self, grupo_id=None, **filtros):
        """
        Genera el reporte de integrantes, filtrando por el grupo indicado.
        """

        def _proceso():
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Integrantes de Grupo"

                font_titulo = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                fill_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

                headers = ["Grupo", "Identificación", "Integrantes (Nombre Completo)", "Email", "Teléfono"]
                ws.append(headers)

                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = font_titulo
                    cell.fill = fill_titulo

                if grupo_id:
                    filtros['grupo_id'] = grupo_id

                queryset = GrupoPersona.objects.filter(**filtros).select_related('grupo', 'persona')

                for gp in queryset:
                    grupo = gp.grupo
                    persona = gp.persona
                    if grupo and persona:
                        ws.append([
                            grupo.nombre,
                            persona.identificacion,
                            persona.nombre_completo_minus(),
                            persona.email,
                            persona.telefono
                        ])

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                nombre_archivo = f"reporte_integrantes_grupo_{self.usuario.id}_{timestamp}.xlsx"
                filepath = os.path.join(self.media_report_path, nombre_archivo)
                wb.save(filepath)

                self._crear_notificacion_exito(
                    titulo="Reporte de Integrantes Listo",
                    descripcion="El reporte de las personas del grupo solicitado ha finalizado.",
                    nombre_archivo=nombre_archivo
                )

            except Exception as e:
                self._crear_notificacion_error(titulo="Reporte de Integrantes", error_msg=e)

        threading.Thread(target=_proceso, daemon=True).start()

    # =========================================================================
    # 5. REPORTE DE SOLICITUDES
    # =========================================================================
    def reporte_solicitudes_thread(self, area_id=None, **filtros):
        """
        Genera reporte de solicitudes. Si recibe area_id, filtra por área;
        caso contrario evalúa otros filtros en **filtros o trae todas.
        """

        def _proceso():
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Solicitudes"

                font_titulo = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                fill_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

                headers = [
                    "ID Solicitud", "Solicitante", "Área", "Descripción",
                    "Estado", "Tipo Solicitud", "Fecha Creación", "Fecha Resolución", "Tiene Evidencia"
                ]
                ws.append(headers)

                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = font_titulo
                    cell.fill = fill_titulo

                if area_id:
                    filtros['area_id'] = area_id

                queryset = Solicitudes.objects.filter(**filtros).select_related('persona', 'area')

                for sol in queryset:
                    ws.append([
                        sol.id,
                        sol.persona.nombre_completo_minus() if sol.persona else "N/A",
                        sol.area.nombre if sol.area else "N/A",
                        sol.descripcion,
                        sol.get_estado_solicitud_display() if hasattr(sol,
                                                                      'get_estado_solicitud_display') else sol.estado_solicitud,
                        sol.get_tipo_solicitud_display() if hasattr(sol,
                                                                    'get_tipo_solicitud_display') else sol.tipo_solicitud,
                        sol.fecha_creacion.strftime('%Y-%m-%d %H:%M') if hasattr(sol,
                                                                                 'fecha_creacion') and sol.fecha_creacion else "N/A",
                        # Asumiendo campo de ModeloBase
                        sol.fecha_resolucion.strftime('%Y-%m-%d %H:%M') if sol.fecha_resolucion else "Pendiente",
                        "Sí" if sol.archivo else "No"
                    ])

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                nombre_archivo = f"reporte_solicitudes_{self.usuario.id}_{timestamp}.xlsx"
                filepath = os.path.join(self.media_report_path, nombre_archivo)
                wb.save(filepath)

                self._crear_notificacion_exito(
                    titulo="Reporte de Solicitudes Listo",
                    descripcion="El informe detallado de solicitudes ya puede ser descargado.",
                    nombre_archivo=nombre_archivo
                )

            except Exception as e:
                self._crear_notificacion_error(titulo="Reporte de Solicitudes", error_msg=e)

        threading.Thread(target=_proceso, daemon=True).start()